import time
import openpyxl
from utils.KeyboardInteraction import inputText

from utils.MouseInteraction import clearName, click_in_square
from ScreenCapture import ScreenCapture
from TextRec import *
from constants.resolution_1920x1080 import *
import cv2
import numpy as np
import random


class ExcelFile:
    def __init__(self, file_path) -> None:
        self.capture = ScreenCapture()
        self.wb = openpyxl.load_workbook(file_path)
        self.ws = self.wb.active
        self.file_path = file_path

    def process_rows(self):
        clearName()
        time.sleep(2)
        self.trashItem = self.getTrashItem()
        previous_values = {
            "col2": None,
            "col3": None,
        }  # Słownik przechowujący poprzednie wartości
        row_number = 2  # Zaczynamy od drugiego rzędu
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            item, value_col2, value_col3, *_ = row  # Rozpakowanie wartości z rzędu
            pos = SEARCH_INPUT_CORDS
            clearName()
            if item == None:
                break
            inputText(
                (pos["posx1"], pos["posy1"]), (pos["posx2"], pos["posy2"]), item.lower()
            )
            if value_col2 != previous_values["col2"]:
                self.switchCategory("tier", value_col2)
                previous_values["col2"] = value_col2

            if value_col3 != previous_values["col3"]:
                self.switchCategory("enchant", value_col3)
                previous_values["col3"] = value_col3

            price = self.scanMarket()
            columnNumber = self.findColumn()
            print(columnNumber)
            self.ws.cell(row=row_number, column=columnNumber, value=price)
            row_number += 1
        self.save()

    def save(self):
        self.wb.save(self.file_path)

    def getTrashItem(self):
        frame = self.capture.get_frame()
        return frame[
            TRASH_ITEM_POS["posy1"] : TRASH_ITEM_POS["posy2"],
            TRASH_ITEM_POS["posx1"] : TRASH_ITEM_POS["posx2"],
        ]

    def scanMarket(self, timeout=random.uniform(2, 2.3)):
        start_time = time.time()
        while True:
            frame = self.capture.get_frame()
            cropped_image = frame[422:444, 1070:1150]
            text = recognize_number(cropped_image)
            if text != "" and (int(text) > 50):
                print("Price:", text)
                return text
            if time.time() - start_time > timeout:
                print("Timeout reached")
                return 0

    def switchCategory(self, category, position):
        if position == None:
            position = 0
        if category == "tier":
            posx1 = TIER_DROPDOWN_CORDS["posx1"]
            posy1 = TIER_DROPDOWN_CORDS["posy2"]
            posx2 = TIER_DROPDOWN_CORDS["posx2"]
            posy2 = TIER_DROPDOWN_CORDS["posy2"]
        elif category == "enchant":
            posx1 = ENCHANT_DROPDOWN_CORDS["posx1"]
            posy1 = ENCHANT_DROPDOWN_CORDS["posy2"]
            posx2 = ENCHANT_DROPDOWN_CORDS["posx2"]
            posy2 = ENCHANT_DROPDOWN_CORDS["posy2"]
        elif category == "quality":
            posx1 = QUALITY_DROPDOWN_CORDS["posx1"]
            posy1 = QUALITY_DROPDOWN_CORDS["posy2"]
            posx2 = QUALITY_DROPDOWN_CORDS["posx2"]
            posy2 = QUALITY_DROPDOWN_CORDS["posy2"]
        print((posx1, posy1), (posx2, posy2))
        click_in_square((posx1, posy1), (posx2, posy2))
        click_in_square(
            (posx1, SUB_ELEMENTS_CORDS[position]["posy1"]),
            (posx2, SUB_ELEMENTS_CORDS[position]["posy2"]),
        )

    def findColumn(self):
        frame = self.capture.get_frame()
        cropped_image = frame[
            AREA_NAME_CORDS["posy1"] : AREA_NAME_CORDS["posy2"],
            AREA_NAME_CORDS["posx1"] : AREA_NAME_CORDS["posx2"],
        ]
        cropped_image = cv2.resize(cropped_image, dsize=None, fx=2, fy=2)
        name = recognize_text(cropped_image)
        name = name.lower().replace("market", "").strip()
        print(name)
        column_index = None
        for col in self.ws.iter_cols(min_row=1, max_row=1, values_only=False):
            cell = col[0]  # Pierwsza komórka w kolumnie (nagłówek)
            if (
                cell.value is not None
                and name is not None
                and cell.value.replace(" ", "") == name.replace(" ", "")
            ):
                column_index = (
                    cell.column
                )  # Znalezienie indeksu kolumny (zwraca literę kolumny)
                return column_index
